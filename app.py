"""
SCOSCHE Packing App — Flask 伺服器
啟動：python app.py
瀏覽器：http://localhost:5001
"""

import json, os, subprocess, sys, threading, tempfile, urllib.request
from datetime import datetime, date as _date
from pathlib import Path
from flask import Flask, abort, jsonify, request, send_file, send_from_directory

BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR / 'data'
JSON_PATH   = DATA_DIR / 'items.json'
TMPL_DIR    = BASE_DIR / 'templates'
STATIC_DIR  = BASE_DIR / 'static'
SCANNER     = BASE_DIR / 'scanner.py'
BUILD_SCRIPT = BASE_DIR / 'scripts' / 'build_packing_list.py'
TEMPLATE_XLSX = BASE_DIR / 'assets' / 'SCOSCHE_PACKING_TEMPLATE.xlsx'

cfg = json.loads((BASE_DIR / 'config.json').read_text(encoding='utf-8')) if (BASE_DIR / 'config.json').exists() else {}
PORT = int(cfg.get('port', 5001))

app = Flask(__name__, template_folder=str(TMPL_DIR), static_folder=str(STATIC_DIR))

# -- DB --
_lock = threading.Lock()
_db: dict = {}

def load_db():
    global _db
    if JSON_PATH.exists():
        raw = json.loads(JSON_PATH.read_text(encoding='utf-8'))
        with _lock:
            _db = {r['item_no']: r for r in raw}
    print(f"[DB] Loaded {len(_db)} items")

def save_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with _lock:
        data = list(_db.values())
    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')


# == Pages ==

@app.route('/')
def home():
    return send_from_directory(str(TMPL_DIR), 'home.html')

@app.route('/items')
def index():
    return send_from_directory(str(TMPL_DIR), 'index.html')

@app.route('/packing')
def packing():
    return send_from_directory(str(TMPL_DIR), 'packing.html')

@app.route('/tools/currency')
def tool_currency():
    return send_from_directory(str(TMPL_DIR), 'currency_v2.html')

@app.route('/tools/cbp')
def tool_cbp():
    # API key 不再注入前端，改由 /api/cbp_analyze 後端代理呼叫 Groq
    return send_from_directory(str(TMPL_DIR), 'cbp_v2.html')

@app.route('/tools/orders')
def tool_orders():
    return send_from_directory(str(TMPL_DIR), 'orders.html')


# == API items ==

@app.route('/api/items', methods=['GET'])
def list_items():
    q = request.args.get('q', '').strip().lower()
    with _lock:
        items = list(_db.values())
    if q:
        items = [it for it in items
                 if q in it.get('item_no', '').lower()
                 or q in it.get('desc', '').lower()]
    result = sorted([{
        'item_no':     it.get('item_no'),
        'desc':        it.get('desc', ''),
        'pcs_per_ctn': it.get('pcs_per_ctn'),
        'dim':         it.get('dim', ''),
        'nw':          it.get('nw'),
        'gw':          it.get('gw'),
        'has_mark':    bool(it.get('shipping_mark', '')),
        'source_file': it.get('source_file', ''),
        'updated_at':  it.get('updated_at', ''),
    } for it in items], key=lambda x: x['item_no'])
    return jsonify(result)


@app.route('/api/items/<item_no>', methods=['GET'])
def get_item(item_no):
    with _lock:
        it = _db.get(item_no)
    if not it:
        abort(404)
    return jsonify(it)


@app.route('/api/items/<item_no>', methods=['PUT'])
def update_item(item_no):
    body = request.get_json(force=True)
    with _lock:
        if item_no not in _db:
            abort(404)
        allowed = {'desc', 'pcs_per_ctn', 'dim', 'nw', 'gw', 'shipping_mark',
                   'has_pallet', 'ctns_per_plt', 'plt_dim', 'plt_nw_extra', 'plt_gw_extra'}
        for k, v in body.items():
            if k in allowed:
                _db[item_no][k] = v
        _db[item_no]['updated_at'] = datetime.now().isoformat()
        # 標記為手動修改過，scanner 增量掃描時不會覆蓋
        _db[item_no]['manual_edited'] = True
    save_db()
    return jsonify({'ok': True})


@app.route('/api/items', methods=['POST'])
def add_item():
    body = request.get_json(force=True)
    item_no = body.get('item_no', '').strip()
    if not item_no:
        return jsonify({'error': 'item_no required'}), 400
    with _lock:
        if item_no in _db:
            return jsonify({'error': f'{item_no} already exists'}), 409
        _db[item_no] = {
            'customer': 'SCOSCHE', 'item_no': item_no,
            'desc': body.get('desc', ''),
            'pcs_per_ctn': body.get('pcs_per_ctn'),
            'dim': body.get('dim', ''),
            'nw': body.get('nw'),
            'gw': body.get('gw'),
            'shipping_mark': body.get('shipping_mark', ''),
            'source_file': 'manual', 'source_mtime': 0,
            'updated_at': datetime.now().isoformat(),
        }
    save_db()
    return jsonify({'ok': True}), 201


@app.route('/api/items/<item_no>', methods=['DELETE'])
def delete_item(item_no):
    with _lock:
        if item_no not in _db:
            abort(404)
        del _db[item_no]
    save_db()
    return jsonify({'ok': True})


@app.route('/api/lookup', methods=['POST'])
def lookup_items():
    body = request.get_json(force=True)
    item_nos = body.get('item_nos', [])
    with _lock:
        result = {no: _db[no] for no in item_nos if no in _db}
    return jsonify(result)


# == API scan ==

_scan_status = {'running': False, 'last': None, 'log': ''}


@app.route('/api/scan', methods=['POST'])
def trigger_scan():
    if _scan_status['running']:
        return jsonify({'error': 'scan in progress'}), 409
    body = request.get_json(force=True, silent=True) or {}
    full_scan = body.get('full', False)

    def run_scan():
        _scan_status['running'] = True
        try:
            args = [sys.executable, str(SCANNER)]
            if full_scan:
                args.append('--full')
            result = subprocess.run(args, capture_output=True, text=True, timeout=180)
            _scan_status['log'] = result.stdout + result.stderr
            load_db()
            _scan_status['last'] = datetime.now().isoformat()
        except Exception as e:
            _scan_status['log'] = str(e)
        finally:
            _scan_status['running'] = False

    threading.Thread(target=run_scan, daemon=True).start()
    return jsonify({'ok': True, 'message': 'scan started'})


@app.route('/api/scan/status', methods=['GET'])
def scan_status():
    return jsonify(_scan_status)


# == API parse email body ==

@app.route('/api/parse_subject', methods=['POST'])
def parse_subject():
    body_req = request.get_json(force=True)
    subj = body_req.get('subject', '').strip()

    from outlook_search import parse_items_from_body, extract_po
    direct_body = body_req.get('body', '').strip()
    if not direct_body:
        return jsonify({
            'outlook_failed': True,
            'error': 'Please paste email body first',
        }), 400
    email_body = direct_body
    email_subj = subj

    po = extract_po(email_body, email_subj)

    with _lock:
        db_snapshot = dict(_db)
    parsed_items = parse_items_from_body(email_body, db=db_snapshot)
    if not parsed_items:
        return jsonify({
            'po': po, 'items': [],
            'email_subject': email_subj,
            'note': 'No items parsed',
            'email_body_preview': email_body[:500],
        })

    import re as _re
    _LASER_RE2  = _re.compile(r'鐳雕號[:：]\s*(\d+)')
    _ITEM_RE2   = _re.compile(r'(8SH[A-Z]\d{6}|S6SH\d{6})', _re.I)
    _lines = email_body.splitlines()
    _laser_sections = []
    for _i, _ln in enumerate(_lines):
        _m = _LASER_RE2.search(_ln)
        if _m:
            _laser_sections.append((_i, _m.group(1)))
    _item_laser_map = {}
    for _i, _ln in enumerate(_lines):
        _m = _ITEM_RE2.search(_ln)
        if _m:
            _ino = _m.group(1)
            _laser = ''
            for _li, _lc in reversed(_laser_sections):
                if _li <= _i:
                    _laser = _lc
                    break
            _item_laser_map[_ino] = _laser

    with _lock:
        db_data = {it['item_no']: _db.get(it['item_no']) for it in parsed_items}

    items_out = []
    for it in parsed_items:
        item_no   = it['item_no']
        db        = db_data.get(item_no)
        ppc       = (db.get('pcs_per_ctn') or 0) if db else 0
        total_pcs = it.get('total_pcs')
        ctns = it.get('ctns')
        if not ctns and total_pcs and ppc:
            ctns = round(total_pcs / ppc)
        date_code = it.get('date_code') or _item_laser_map.get(item_no, '')
        items_out.append({
            'item_no':       item_no,
            'desc':          db.get('desc', '')          if db else '',
            'ctns':          ctns,
            'total_pcs':     total_pcs,
            'pcs_per_ctn':   ppc or None,
            'dim':           db.get('dim', '')            if db else '',
            'nw':            db.get('nw')                 if db else None,
            'gw':            db.get('gw')                 if db else None,
            'shipping_mark': db.get('shipping_mark', '')  if db else '',
            'date_code':     date_code,
            'rfid':          False,
            'db_found':      bool(db),
        })

    return jsonify({
        'po':           po,
        'items':        items_out,
        'email_subject': email_subj,
    })


@app.route('/api/ship_to', methods=['GET'])
def get_ship_to():
    ship_to_file = DATA_DIR / 'ship_to.json'
    defaults = [
        "Scosche Industries, Inc.\n1550 Pacific Ave\nOxnard, CA 93033\nUSA",
        "Scosche Industries, Inc.\n201 Quality Circle\nHuntsville, AL 35806\nUSA",
    ]
    if ship_to_file.exists():
        return jsonify(json.loads(ship_to_file.read_text(encoding='utf-8')))
    return jsonify(defaults)


@app.route('/api/ship_to', methods=['POST'])
def save_ship_to():
    data = request.get_json(force=True)
    ship_to_file = DATA_DIR / 'ship_to.json'
    ship_to_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    return jsonify({'ok': True})


# == API generate Excel ==

@app.route('/api/generate_excel', methods=['POST'])
def generate_excel():
    body        = request.get_json(force=True)
    po          = body.get('po', '').strip()
    # 前端沒填日期時退回今天，避免 Excel 的 PRINT AT 空白
    ship_date   = body.get('ship_date', '').strip() or _date.today().isoformat()
    payment     = body.get('payment_term', '').strip() or 'T/T NET ETD 75DAYS'
    ship_to     = body.get('ship_to', '').strip()
    items       = body.get('items', [])

    if not po:
        return jsonify({'error': 'PO# required'}), 400
    if not items:
        return jsonify({'error': 'No items'}), 400
    if not BUILD_SCRIPT.exists():
        return jsonify({'error': f'Build script not found: {BUILD_SCRIPT}'}), 500
    if not TEMPLATE_XLSX.exists():
        return jsonify({'error': f'Template not found: {TEMPLATE_XLSX}'}), 500

    # Auto-fill shipping_mark and pallet fields from DB
    with _lock:
        for it in items:
            db = _db.get(it.get('item_no', ''))
            if db:
                if not it.get('shipping_mark'):
                    it['shipping_mark'] = db.get('shipping_mark', '')
                # Pull pallet fields from DB if item has them and caller didn't specify
                if not it.get('has_pallet') and db.get('has_pallet'):
                    it['has_pallet']   = True
                    it['ctns_per_plt'] = db.get('ctns_per_plt')
                    it['plt_dim']      = db.get('plt_dim', '')
                    it['plt_nw_extra'] = db.get('plt_nw_extra', 0)
                    it['plt_gw_extra'] = db.get('plt_gw_extra', 0)

    out_path = str(DATA_DIR / f'packing_PO{po}.xlsx')

    items_config = []
    for it in items:
        item_cfg = {
            'item_no':       it.get('item_no', ''),
            'desc':          it.get('desc', ''),
            'ctns':          int(it.get('ctns') or 0),
            'pcs_per_ctn':   int(it.get('pcs_per_ctn') or 0),
            'dim':           it.get('dim', ''),
            'nw':            float(it.get('nw') or 0),
            'gw':            float(it.get('gw') or 0),
            'date_code':     it.get('date_code', ''),
            'rfid':          bool(it.get('rfid', False)),
            'shipping_mark': it.get('shipping_mark', ''),
        }
        if it.get('has_pallet'):
            item_cfg['has_pallet']   = True
            item_cfg['ctns_per_plt'] = it.get('ctns_per_plt')
            item_cfg['plt_dim']      = it.get('plt_dim', '')
            item_cfg['plt_nw_extra'] = float(it.get('plt_nw_extra') or 0)
            item_cfg['plt_gw_extra'] = float(it.get('plt_gw_extra') or 0)
        items_config.append(item_cfg)

    config = {
        'po': po, 'date': ship_date, 'payment_term': payment,
        'ship_to': ship_to,
        'template_path': str(TEMPLATE_XLSX),
        'output_path': out_path,
        'items': items_config,
    }

    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        cfg_path = f.name

    try:
        result = subprocess.run(
            [sys.executable, str(BUILD_SCRIPT), '--config', cfg_path],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            return jsonify({'error': (result.stderr or result.stdout or 'Build failed')[:500]}), 500
        if not os.path.exists(out_path):
            return jsonify({'error': 'Output file missing'}), 500

        return send_file(out_path, as_attachment=True,
                         download_name=f'SCOSCHE PACKING PO#{po}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        try:
            os.unlink(cfg_path)
        except Exception:
            pass


# == API CBP Analyze (Groq proxy) ==

@app.route('/api/cbp_analyze', methods=['POST'])
def cbp_analyze():
    groq_key = cfg.get('groq_api_key', '').strip()
    if not groq_key:
        return jsonify({'error': 'Please set groq_api_key in config.json'}), 500

    body = request.get_json(force=True)
    payload = json.dumps({
        'model': 'meta-llama/llama-4-scout-17b-16e-instruct',
        'messages': body.get('messages', []),
        'temperature': body.get('temperature', 0.3),
        'max_tokens': body.get('max_tokens', 1024),
    }).encode('utf-8')

    req = urllib.request.Request(
        'https://api.groq.com/openai/v1/chat/completions',
        data=payload,
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {groq_key}',
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode('utf-8'))
        return jsonify(result)
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')
        return jsonify({'error': f'Groq API error {e.code}', 'detail': err_body}), e.code
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# == API filter orders (openpyxl, preserve formatting) ==

@app.route('/api/filter_orders', methods=['POST'])
def filter_orders():
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        return jsonify({'error': 'Missing openpyxl'}), 500

    f = request.files.get('file')
    start_str = request.form.get('start', '')
    end_str   = request.form.get('end', '')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    if not start_str or not end_str:
        return jsonify({'error': 'Please provide start and end dates'}), 400
    try:
        start_d = _date.fromisoformat(start_str)
        end_d   = _date.fromisoformat(end_str)
    except ValueError:
        return jsonify({'error': 'Date format error'}), 400

    fname = f.filename or ''
    is_xls = fname.lower().endswith('.xls') and not fname.lower().endswith('.xlsx')
    suffix = '.xls' if is_xls else '.xlsx'

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        if is_xls:
            try:
                import xlrd
            except ImportError:
                return jsonify({'error': 'Missing xlrd, run: pip install xlrd'}), 500
            xls = xlrd.open_workbook(tmp_path)
            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name in xls.sheet_names():
                xls_ws = xls.sheet_by_name(sheet_name)
                new_ws = wb.create_sheet(title=sheet_name)
                for ri in range(xls_ws.nrows):
                    for ci in range(xls_ws.ncols):
                        cell = xls_ws.cell(ri, ci)
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                dt_t = xlrd.xldate_as_tuple(cell.value, xls.datemode)
                                val = _date(*dt_t[:3]) if dt_t[3:] == (0,0,0) else datetime(*dt_t)
                            except Exception:
                                val = cell.value
                        else:
                            val = cell.value
                        new_ws.cell(ri + 1, ci + 1, val)
        else:
            wb = load_workbook(tmp_path)

        ws = None
        for sn in wb.sheetnames:
            if 'SCOSCHE' in sn.upper():
                ws = wb[sn]
                break
        if ws is None:
            ws = wb.active

        # Find date column (header on row 2, data from row 4)
        header_row = 2
        date_col = None
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(header_row, col).value or '').lower()
            if 'ready' in h or 'shipment' in h:
                date_col = col
                break
        if date_col is None:
            date_col = 8

        rows_to_delete = []
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row, date_col).value
            if val is None:
                rows_to_delete.append(row)
                continue
            if isinstance(val, (datetime, _date)):
                d = val.date() if isinstance(val, datetime) else val
            else:
                try:
                    d = _date.fromisoformat(str(val).strip()[:10])
                except Exception:
                    rows_to_delete.append(row)
                    continue
            if not (start_d <= d <= end_d):
                rows_to_delete.append(row)

        for row in reversed(rows_to_delete):
            ws.delete_rows(row)

        kept = max(ws.max_row - 3, 0)

        # 存到記憶體再回傳，避免暫存檔堆積
        import io as _io
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = send_file(
            buf,
            as_attachment=True,
            download_name='filtered_orders.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp.headers['X-Filtered-Count'] = str(kept)
        return resp

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


 
if __name__ == '__main__':
    load_db()
    host = '127.0.0.1'   # 僅限本機使用；如需區網共用請自行改回 0.0.0.0（注意 API 無驗證）
    print('=' * 50)
    print(f'SCOSCHE Packing App — http://localhost:{PORT}')
    print(f'Home:     http://localhost:{PORT}/')
    print(f'Items:    http://localhost:{PORT}/items')
    print(f'Packing:  http://localhost:{PORT}/packing')
    print(f'Currency: http://localhost:{PORT}/tools/currency')
    print(f'CBP:      http://localhost:{PORT}/tools/cbp')
    print(f'Orders:   http://localhost:{PORT}/tools/orders')
    print('=' * 50)
    app.run(host=host, port=PORT, debug=False)
