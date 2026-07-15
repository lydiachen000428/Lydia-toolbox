#!/usr/bin/env python3
"""
SCOSCHE Packing List Builder
Generates a formatted Excel packing list (.xlsx) from a JSON config file.

Usage:
    python build_packing_list.py --config /path/to/config.json

Config JSON schema:
{
  "po": "163031",
  "date": "2026-06-24",
  "payment_term": "T/T NET ETD 75DAYS",
  "template_path": "/path/to/SCOSCHE_PACKING_TEMPLATE.xlsx",
  "output_path": "/path/to/output/SCOSCHE PACKING PO#163031.xlsx",
  "items": [
    {
      "item_no": "8SHB130193",
      "desc": "0266MP2WD-SP1; MAGICMOUNT PRO WIN DASH",
      "ctns": 79,
      "pcs_per_ctn": 40,
      "dim": "55*47.5*20.5",    <- W*D*H in cm, * as separator
      "nw": 4.70,
      "gw": 8.90,
      "rfid": true,
      "date_code": "1630313226" <- 鐳雕號 from PO email
    }
  ]
}

Notes:
- dim format: "W*D*H" in cm matching template formula style
- Each item CTN range always starts from 1
- Shipping marks use the standard SCOSCHE format (Part Number / PO# / Date Code / C/N / Qty / NW / GW / Dims mm / Dims inches / Made in Vietnam / ROHS)
- Items 1-3 go across 3 columns (A/E/I); items 4+ continue below in col A
- Pallet items (has_pallet=true) trigger automatic generation of a second "WITH PLT" sheet.
  Extra fields required per pallet item:
    "has_pallet":    true,
    "ctns_per_plt":  240,          <- CTNs per pallet
    "plt_dim":       "121.92*101.6*71.2",  <- loaded pallet W*D*H in cm
    "plt_nw_extra":  9.92,         <- extra KG added to N.W. w/pallet (pallet tare)
    "plt_gw_extra":  4.0           <- extra KG added to G.W. w/pallet
  Optional: "po" per item overrides global po on WITH PLT sheet.
"""

import argparse
import tempfile, copy, json, math, os, shutil, sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# Red font: MUST be FFFF0000 (fully opaque). 'FF0000' = transparent (00FF0000).
RED_FONT = Font(name='Times New Roman', color='FFFF0000')

# ── WITH PLT sheet constants ──────────────────────────────────────────────────
_PLT_HEADER_A = [
    'COMART CORPORATION',
    'TEL:+886-2-89111133   FAX:+886-2-86653056',
    '3F., No.12, Lane 235, Baociao Rd., Sindian Dist, New Taipei City 23145, Taiwan, R.O.C.',
    'PACKING LIST',
    'CONSIGNEE: SCOSCHE INDUSTRIES INC.',
    'SHIP TO: 188 COMMERCE WAY SPRUCE PINE, AL 35585, USA',
    'ATTN: DUSTIN MURRAY',
    'EIN: 95-3659275',
    'TEL: 256-277-0045',
]

_PLT_COL_HEADERS = [
    (1,  'CTN #'),
    (4,  '# of CTNS'),
    (5,  '# of Pallets'),
    (9,  'PO#'),
    (10, 'Item No'),
    (11, 'Description'),
    (12, 'Quantity'),
    (13, 'CTN/PLT'),
    (14, 'N.W.(CTN)'),
    (15, 'G.W. (CTN)'),
    (16, 'Dimension(CM)'),
    (17, 'CBM W/Pallet'),
    (18, 'N.W. W/Pallet'),
    (19, 'G.W.W/Pallet'),
]

KG_TO_LBS = 2.20462
MM_TO_IN   = 0.03937


def copy_cell_style(src_cell, dst_cell):
    """Copy all formatting from src to dst."""
    if src_cell.has_style:
        dst_cell.font          = copy.copy(src_cell.font)
        dst_cell.border        = copy.copy(src_cell.border)
        dst_cell.fill          = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment     = copy.copy(src_cell.alignment)


def write_shipping_mark(ws, row_start, col, item, po):
    """
    Write one shipping mark block.
    Strategy A: use the stored shipping_mark from DB as the base template —
    keeps Part Number / Dims / Made in Vietnam / ROHS exactly as they appear
    in historical files.  Only PO#, Date Code, C/N, Qty, NW, GW are replaced
    with per-shipment values.

    Fallback (item not in DB / no stored mark): generate dynamically,
    using only the portion of desc before the semicolon as the Part Number.

    Returns the row AFTER the last written line.
    """
    nw       = item.get('nw', 0) or 0
    gw       = item.get('gw', 0) or 0
    nw_lb    = round(nw * KG_TO_LBS, 2)
    gw_lb    = round(gw * KG_TO_LBS, 2)
    date_code = item.get('date_code', '') or '待確認'
    ctns     = item.get('ctns', 0)
    ppc      = item.get('pcs_per_ctn', 0)

    stored_mark = (item.get('shipping_mark') or '').strip()

    if stored_mark:
        # Use stored mark; replace only the per-shipment dynamic fields
        lines = []
        for ln in stored_mark.splitlines():
            low = ln.lower().lstrip()
            if low.startswith('po#:'):
                lines.append(f"PO#: {po}")
            elif low.startswith('date code:'):
                lines.append(f"Date Code: {date_code}")
            elif low.startswith('c/n:'):
                lines.append(f"C/N: of {ctns}")
            elif low.startswith('qty:'):
                lines.append(f"Qty: {ppc} pcs")
            elif low.startswith('n. w.:') or low.startswith('n.w.:'):
                lines.append(f"N. W.: {nw} kg {nw_lb} lbs")
            elif low.startswith('g. w.:') or low.startswith('g.w.:'):
                lines.append(f"G. W.: {gw} kg {gw_lb} lbs")
            else:
                lines.append(ln)   # Part Number / Dims / Vietnam / ROHS 原封不動
    else:
        # Fallback: generate dynamically (item not yet in DB)
        part_no = (item.get('desc') or '').split(';')[0].strip() or item.get('item_no', '')
        try:
            w_cm, d_cm, h_cm = [float(x) for x in item['dim'].split('*')]
            w_mm = int(round(w_cm * 10))
            d_mm = int(round(d_cm * 10))
            h_mm = int(round(h_cm * 10))
            w_in = round(w_mm * MM_TO_IN, 1)
            d_in = round(d_mm * MM_TO_IN, 1)
            h_in = round(h_mm * MM_TO_IN, 1)
            dim_mm = f"{w_mm} X {d_mm} X {h_mm} mm"
            dim_in = f"{w_in} X {d_in} X {h_in} inches"
        except Exception:
            dim_mm = item.get('dim', '')
            dim_in = ''
        lines = [
            f"Part Number: {part_no}",
            f"PO#: {po}",
            f"Date Code: {date_code}",
            f"C/N: of {ctns}",
            f"Qty: {ppc} pcs",
            f"N. W.: {nw} kg {nw_lb} lbs",
            f"G. W.: {gw} kg {gw_lb} lbs",
            f"Dims (WDH): {dim_mm}",
            f"(WDH): {dim_in}",
            "Made in Vietnam",
            "*** GOODS ROHS COMPLIANT***",
        ]

    if item.get('rfid', False):
        lines.append("RFID LABEL REQUIRED")

    for i, val in enumerate(lines):
        cell = ws.cell(row=row_start + i, column=col)
        cell.value = val
        cell.font  = RED_FONT

    return row_start + len(lines)


def build_with_plt_sheet(wb, po, today, pallet_items):
    """
    Add 'WITH PLT' sheet for palletised items.
    pallet_items: items from config that have has_pallet=True.
    Each item must include: ctns, pcs_per_ctn, nw, gw, desc, item_no,
                            ctns_per_plt, plt_dim, plt_nw_extra, plt_gw_extra.
    Optional per-item 'po' key overrides the global po on this sheet.
    """
    ws = wb.create_sheet('WITH PLT')

    # ── Company header (rows 1-10) ────────────────────────────────────────
    for r, text in enumerate(_PLT_HEADER_A, start=1):
        ws.cell(row=r, column=1).value = text

    ws['Q7']  = 'INV: '
    ws['Q8']  = f'PO: {po}'
    ws['Q9']  = f'PRINT AT: {today}'
    ws['A10'] = 'PAYMENT TERM: T/T NET ETD 75DAYS'
    ws['Q10'] = 'CURRENCY: USD'

    # ── Column headers (row 11) ──────────────────────────────────────────
    for col, label in _PLT_COL_HEADERS:
        ws.cell(row=11, column=col).value = label

    # ── Data rows (2 rows per item: data + subtotal) ─────────────────────
    n             = len(pallet_items)
    first_dr      = 12
    data_rows     = []
    subtotal_rows = []

    for idx, item in enumerate(pallet_items):
        dr = first_dr + idx * 2
        sr = dr + 1
        data_rows.append(dr)
        subtotal_rows.append(sr)

        item_po      = item.get('po') or po
        ctns_per_plt = item.get('ctns_per_plt', 1)
        plt_dim      = item.get('plt_dim', '')
        plt_nw_extra = item.get('plt_nw_extra') or 0
        plt_gw_extra = item.get('plt_gw_extra') or 0

        # Data row
        ws.cell(dr, 1).value  = 1
        ws.cell(dr, 2).value  = '-'
        ws.cell(dr, 3).value  = item['ctns']
        ws.cell(dr, 4).value  = f'=C{dr}-A{dr}+1'       # # of CTNs
        ws.cell(dr, 5).value  = f'=D{dr}/M{dr}'          # # of Pallets
        ws.cell(dr, 9).value  = item_po
        ws.cell(dr, 10).value = item['item_no']
        ws.cell(dr, 11).value = item.get('desc', '')
        ws.cell(dr, 12).value = item.get('pcs_per_ctn', '')
        ws.cell(dr, 13).value = ctns_per_plt
        ws.cell(dr, 14).value = item.get('nw', '')
        ws.cell(dr, 15).value = item.get('gw', '')
        ws.cell(dr, 16).value = plt_dim

        # CBM w/pallet: ={plt_dim}/1000000  (e.g. =121.92*101.6*71.2/1000000)
        ws.cell(dr, 17).value = f'={plt_dim}/1000000'

        def _num(v):
            """Format extra weight: int if whole number, else float."""
            return int(v) if v == int(v) else v

        # N.W. w/pallet: =N{dr}*M{dr} [+extra]
        nw_formula = f'=N{dr}*M{dr}'
        if plt_nw_extra:
            nw_formula += f'+{_num(plt_nw_extra)}'
        ws.cell(dr, 18).value = nw_formula

        # G.W. w/pallet: =O{dr}*M{dr} [+extra]
        gw_formula = f'=O{dr}*M{dr}'
        if plt_gw_extra:
            gw_formula += f'+{_num(plt_gw_extra)}'
        ws.cell(dr, 19).value = gw_formula

        # Subtotal row
        ws.cell(sr, 11).value = 'SUBTOTAL:'
        ws.cell(sr, 12).value = f'=L{dr}*D{dr}'
        ws.cell(sr, 17).value = f'=Q{dr}*E{dr}'
        ws.cell(sr, 18).value = f'=R{dr}*E{dr}'
        ws.cell(sr, 19).value = f'=S{dr}*E{dr}'

    # ── TOTAL row ────────────────────────────────────────────────────────
    total_row = first_dr + n * 2
    last_dr   = data_rows[-1]

    ws.cell(total_row, 4).value  = f'=SUM(D{first_dr}:D{last_dr})'
    ws.cell(total_row, 5).value  = f'=SUM(E{first_dr}:E{last_dr})'
    ws.cell(total_row, 11).value = 'TOTAL:'
    ws.cell(total_row, 12).value = '=' + '+'.join(f'L{r}' for r in subtotal_rows)
    ws.cell(total_row, 17).value = '=' + '+'.join(f'Q{r}' for r in subtotal_rows)
    ws.cell(total_row, 18).value = '=' + '+'.join(f'R{r}' for r in subtotal_rows)
    ws.cell(total_row, 19).value = '=' + '+'.join(f'S{r}' for r in subtotal_rows)


def build_packing_list(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    po           = cfg['po']
    _today_raw   = cfg.get('date', str(date.today()))
    try:
        from datetime import datetime as _dt
        _d = _dt.strptime(_today_raw, '%Y-%m-%d')
        today = f"{_d.day} {_d.strftime('%B').upper()} {_d.year}"  # e.g. 8 JULY 2026
    except Exception:
        today = _today_raw
    payment_term = 'T/T NET ETD 75DAYS'  # 固定值，不從 config 讀
    items        = cfg['items']
    n            = len(items)
    template     = cfg['template_path']
    output       = cfg['output_path']

    # Work on a temp copy
    tmp = os.path.join(tempfile.gettempdir(), f'PKG_{po}_build.xlsx')
    shutil.copy2(template, tmp)
    os.chmod(tmp, 0o644)

    # Load template separately for style reference
    tmpl_wb = load_workbook(template, data_only=False)
    tmpl_ws = tmpl_wb['PAC']

    wb = load_workbook(tmp, data_only=False)
    ws = wb['PAC']

    # ── Headers ──────────────────────────────────────────────────────────
    # J7 (INV) intentionally left blank — user fills manually
    ws['J8'].value  = f'PO: {po}'
    ws['J9'].value  = f'PRINT AT: {today}'
    ws['A10'].value = f'PAYMENT TERM: {payment_term}'

    # ── Row layout ───────────────────────────────────────────────────────
    extra_rows = (n - 1) * 2
    total_row  = 12 + n * 2

    if extra_rows > 0:
        ws.insert_rows(14, extra_rows)

        for mc in list(ws.merged_cells.ranges):
            if str(mc) == 'A14:C14':
                ws.merged_cells.ranges.remove(mc)
                ws.merge_cells(f'A{total_row}:C{total_row}')
                break

        for i in range(1, n):
            data_row = 12 + i * 2
            sub_row  = data_row + 1
            for col in range(1, 13):
                copy_cell_style(tmpl_ws.cell(12, col), ws.cell(data_row, col))
                copy_cell_style(tmpl_ws.cell(13, col), ws.cell(sub_row, col))

    # ── Data rows ────────────────────────────────────────────────────────
    total_ctns = sum(item['ctns'] for item in items)

    for idx, item in enumerate(items):
        dr  = 12 + idx * 2
        sr  = dr + 1
        dim = item['dim']

        ws[f'A{dr}'].value = '1'
        ws[f'B{dr}'].value = '-'
        ws[f'C{dr}'].value = str(item['ctns'])
        ws[f'D{dr}'].value = f'=C{dr}-A{dr}+1'
        ws[f'E{dr}'].value = po
        ws[f'F{dr}'].value = item['item_no']
        ws[f'G{dr}'].value = item.get('desc', '')
        ws[f'H{dr}'].value = item.get('pcs_per_ctn', '')
        ws[f'I{dr}'].value = dim
        ws[f'J{dr}'].value = f'={dim}/1000000'
        ws[f'K{dr}'].value = item.get('nw', '')
        ws[f'L{dr}'].value = item.get('gw', '')

        ws[f'G{sr}'].value = 'SUBTOTAL:'
        ws[f'H{sr}'].value = f'=H{dr}*D{dr}'
        ws[f'J{sr}'].value = f'=J{dr}*D{dr}'
        ws[f'K{sr}'].value = f'=K{dr}*D{dr}'
        ws[f'L{sr}'].value = f'=L{dr}*D{dr}'

    # ── TOTAL row ────────────────────────────────────────────────────────
    sub_rows = [str(13 + i * 2) for i in range(n)]
    ws[f'A{total_row}'].value = f'{total_ctns} CTNS'
    ws[f'G{total_row}'].value = 'TOTAL:'
    ws[f'H{total_row}'].value = '=' + '+'.join(f'H{r}' for r in sub_rows)
    ws[f'J{total_row}'].value = '=' + '+'.join(f'J{r}' for r in sub_rows)
    ws[f'K{total_row}'].value = '=' + '+'.join(f'K{r}' for r in sub_rows)
    ws[f'L{total_row}'].value = '=' + '+'.join(f'L{r}' for r in sub_rows)

    # ── Clear residual template mark content ────────────────────────────────
    # Template PAC pre-fills a single mark block at rows 18-29 (col A only).
    # After row insertion it shifts to rows (18+extra_rows) through (29+extra_rows).
    # New marks overwrite most of it, but any tail lines (e.g. the RFID on
    # template row 29) survive when the new mark is shorter. Clear them first.
    for _r in range(18 + extra_rows, 30 + extra_rows):
        ws.cell(_r, 1).value = None
    print('[DBG] A45 after clear:', repr(ws.cell(45,1).value))

    # ── Shipping marks ────────────────────────────────────────────────────
    # Layout: groups of 3, placed at cols A / G / I.
    # Each group on the same row set; 1 blank row between groups.
    # Last group with fewer than 3 items: unused slots stay blank.
    MARK_START = total_row + 4
    MARK_COLS  = [1, 7, 9]   # A, G, I

    cur_row = MARK_START
    for grp in range(0, n, 3):
        group = items[grp:grp + 3]
        group_height = max(
            12 if item.get('rfid', False) else 11
            for item in group
        )
        for col_idx, item in enumerate(group):
            write_shipping_mark(ws, cur_row, MARK_COLS[col_idx], item, po)
        cur_row += group_height + 1  # +1 blank row between groups

    # ── WITH PLT sheet (generated when any item has pallet data) ─────────
    pallet_items = [item for item in items if item.get('has_pallet')]
    if pallet_items:
        build_with_plt_sheet(wb, po, today, pallet_items)
        print(f'  WITH PLT sheet: {len(pallet_items)} item(s)')

    # -- Save --